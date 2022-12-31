from pdflatex import PDFLaTeX
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import re
import generateAnswerKey


def hyphenateText(text):
    newText = text.replace(' ', '-')
    return newText


def deHyphenateText(text):
    newText = text.replace('-', ' ')
    return newText

spreadsheet = openpyxl.load_workbook('../sections.xlsx')

for i in range(0, len(spreadsheet.sheetnames)):
    unitName = hyphenateText(spreadsheet.sheetnames[i])
    unitNameUnhyphenated = deHyphenateText(unitName)
    # try:
    #     with open(
    #             f"{unitName}-content.tex", "x") as reviewExercisesLaTeX:
    #         reviewExercisesLaTeX.write('')
    # except FileExistsError:
    #     print(f"File {unitName} already exists; skipping.")

    with open(f'content/{unitName}-content.tex',
              encoding='utf-8') as file:
        appendContent = file.read()
        numProblems = appendContent.count("\\begin{question}")
        templateCopy = f"review-exercises-{unitName}.tex"   # the file for final pdf
        shutil.copyfile(f'content/review-exercises-template.tex', templateCopy)

        file.close()

    with open(templateCopy, encoding='utf-8') as file:
        data = file.read()
        data = data.replace('[[[APPEND QUESTIONS HERE]]]', appendContent).\
            replace('[[[NUMBER OF PROBLEMS]]]', str(numProblems)).\
            replace('[[[CHAPTER NAME]]]', unitNameUnhyphenated).\
            replace('[[[APPEND ANSWERS HERE]]]', str(generateAnswerKey.generateAnswerKey(f'content/{unitName}-content.tex')))
        print(generateAnswerKey.generateAnswerKey(f'content/{unitName}-content.tex'))
    with open(templateCopy, 'w', encoding='utf-8') as file:
        file.write(data)
        file.close()

    os.system(f"pdflatex {templateCopy}")
