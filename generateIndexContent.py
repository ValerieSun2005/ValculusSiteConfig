import openpyxl
from bs4 import BeautifulSoup
import dominate
from dominate.tags import *
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


spreadsheet = openpyxl.load_workbook('sections.xlsx')
doc = dominate.document(title='Sections')

def hyphenateText(text):
    newText = text.replace(' ', '-')
    return newText

def indexContent():
    appendText = '<link href="css/section.css" rel="stylesheet">\n' \
                 '<link href="css/layout.css" rel="stylesheet">\n' \
                 '<link href="css/layout-print.css" media="print" rel="stylesheet">'


    for i in range(0, len(spreadsheet.sheetnames)):
        sectionSheet = spreadsheet[f'{spreadsheet.sheetnames[i]}']
        # here i = unitNumber
        if i > 0:
            appendText += '<hr> \n'
        else:
            pass
        appendText += f'<h3>{str(i)} — {spreadsheet.sheetnames[i]}</h3>\n'

        appendText += f'<div class="section-unit">\n' \
                      f'<div class="sections">\n' \
                      f'<ul>\n'
        for j in range(1, sectionSheet.max_row + 1):
            appendText += '<li>\n' \
                          f'<a href="calculus/{i}.{j}-{hyphenateText(sectionSheet.cell(row = j, column = 1).value)}.html">' \
                          f'{i}.{j} — {sectionSheet.cell(row = j, column = 1).value}' \
                          f'</a>\n' \
                          f'<span class="section-info">ⓘ</span>' \
                          '</li>'
        appendText += f'</ul>\n' \
                      f'</div>\n' \
                      f'<img class="header-image" src="images/indexImages/{i}.jpg" alt="">' \
                      f'</div>\n'
    return appendText
