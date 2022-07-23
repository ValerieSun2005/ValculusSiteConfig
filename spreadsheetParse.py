import openpyxl
from bs4 import BeautifulSoup
import dominate
from dominate.tags import *
import re
from openpyxl import Workbook
from openpyxl import load_workbook
# import sectionCreate
import pandas as pd

#
# generates spreadsheet. warning: overrides
# def createSpreadsheet():
#     wb = Workbook()
#
#     # path of spreadsheet of section labels and descriptions
#     filepath = 'sections.xlsx'
#
#     wb.save(filepath)
#
#     # load spreadsheet
#     wb = load_workbook(filepath)
#     wb.remove(wb['Sheet'])
#     for i in range(0, len(sectionCreate.topicsList)):
#         wb.create_sheet(f'Unit{i}')  # create new sheet in file
#         sectionSheet = wb[f'Unit{i}']
#         for j in range(0, len(sectionCreate.topicsList[i])):
#             sectionSheet[f'A{j + 1}'] = sectionCreate.topicsList[i][j]
#             try:
#                 sectionSheet[f'B{j + 1}'] = sectionCreate.topicsInfo[i][j]
#             except:
#                 pass
#
#         wb.save(filepath)
#
# sheet = workbook.active


spreadsheet = openpyxl.load_workbook('sections.xlsx')
# print(spreadsheet.sheetnames)


for i in range(0, len(spreadsheet.sheetnames)):
    sectionSheet = spreadsheet[f'{spreadsheet.sheetnames[i]}']
    for j in range(1, sectionSheet.max_row + 1):
        print(sectionSheet.cell(row = j, column = 1).value)

# cell_obj = employees_sheet.cell(row=1, column=1)