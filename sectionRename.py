import os
from bs4 import BeautifulSoup
import dominate
from dominate.tags import *
import re
import pageCreate
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

spreadsheet = openpyxl.load_workbook('sections.xlsx')

def hyphenateText(text):
    newText = text.replace(' ', '-')
    return newText

def rename():
    unitOfRenamedSection = input('In What Unit Is the Section You Want To Rename? Unit: ')

    try:
        unitOfRenamedSection = int(unitOfRenamedSection)
        sheetRenamedSection = spreadsheet[
            f'{spreadsheet.sheetnames[unitOfRenamedSection]}']  # finds sheet of unit to rename
        oldSectionNameInput = input('Old Section Name (Case Insensitive): ').lower()
        for i in range(1, sheetRenamedSection.max_row + 1):
            if sheetRenamedSection.cell(row=i, column=1).value.lower() == oldSectionNameInput:  # searches all the rows
                # in column A
                newSectionName = input('New Section Name (Case Sensitive): ')
                oldSectionName = sheetRenamedSection[f'A{i}'].value  # gets old section name, before it is overwritten
                sheetRenamedSection[f'A{i}'] = newSectionName  # overwrites section name
                spreadsheet.save(filename = 'sections.xlsx')  # saves changes. writing done

                # renames sectionMaterials/{unit}
                os.rename(f"sectionMaterials/{oldSectionName}",
                          f'sectionMaterials/{newSectionName}')

                # renames all the html content in the sectionMaterials/{unit} folder
                os.rename(f"sectionMaterials/{newSectionName}/content-{hyphenateText(oldSectionName)}.html",
                          f"sectionMaterials/{newSectionName}/content-{hyphenateText(newSectionName)}.html")
                os.rename(f"sectionMaterials/{newSectionName}/exerciseList-{hyphenateText(oldSectionName)}.html",
                          f"sectionMaterials/{newSectionName}/exerciseList-{hyphenateText(newSectionName)}.html")

                # all done! <3
                print(f"New Section: {unitOfRenamedSection}.{i}: {sheetRenamedSection[f'A{i}'].value}")
                print('Success! :)')

                path = 'Valculus/calculus'
                for k in os.listdir(path):
                    os.remove(os.path.join(path, k))

                pageCreate.create_all_pages(f'{unitOfRenamedSection}.{i}',
                                            f'{sheetRenamedSection[f"A{i}"].value}')
            else:
                pass


    except IndexError:
        print(f"Error: '{unitOfRenamedSection}' is not between 0 and 10")
    except ValueError:
        print(f"Error: '{unitOfRenamedSection}' is not an integer")


    import buildSectionPage
    buildSectionPage.generateContentPages()
    # buildSectionPage.generateLabeledSections()
rename()