import openpyxl
from bs4 import BeautifulSoup
import dominate
from dominate.tags import *
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pageCreate
import sectionCrossReferenceManager
import pandas as pd


spreadsheet = openpyxl.load_workbook('sections.xlsx')
doc = dominate.document(title='Sections')

def hyphenateText(text):
    newText = text.replace(' ', '-')
    return newText

def generateContentPages():
    for i in range(0, len(spreadsheet.sheetnames)):
        sectionSheet = spreadsheet[f'{spreadsheet.sheetnames[i]}']
        for j in range(1, sectionSheet.max_row + 1):
            sectionLabelHyphenated = hyphenateText(
                sectionSheet.cell(row=j, column=1).value)  # convert spaces in file name to hyphens
            try:
                os.mkdir(f"sectionMaterials/{sectionSheet.cell(row=j, column=1).value}")
                with open(
                        f"sectionMaterials/{sectionSheet.cell(row=j, column=1).value}/content-{sectionLabelHyphenated}.html",
                        'x') as htmlContent:
                    htmlContent.write('')
                with open(
                        f"sectionMaterials/{sectionSheet.cell(row=j, column=1).value}/exerciseList-{sectionLabelHyphenated}.html",
                        'x') as htmlExerciseList:
                    htmlExerciseList.write('')
            except FileExistsError:
                pass

def generateLabeledSections():
    path = 'Valculus/calculus'
    for k in os.listdir(path):
        os.remove(os.path.join(path, k))
    for i in range(0, len(spreadsheet.sheetnames)):
        sectionSheet = spreadsheet[f'{spreadsheet.sheetnames[i]}']
        for j in range(1, sectionSheet.max_row + 1):
            pageCreate.create_all_pages(f'{i}.{j}', sectionSheet.cell(row=j, column=1).value)

generateLabeledSections()
# sectionCrossReferenceManager.tagSectionMaterials()
sectionCrossReferenceManager.insertReferences()
