import os
from bs4 import BeautifulSoup
import dominate
from dominate.tags import *
import re
import pageCreate
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

spreadsheet = openpyxl.load_workbook('sections.xlsx')

def hyphenateText(text):
    newText = text.replace(' ', '-')
    return newText

for i in range(0, len(spreadsheet.sheetnames)):
    sectionSheet = spreadsheet[f'{spreadsheet.sheetnames[i]}']
    for j in range(1, sectionSheet.max_row + 1):
        sectionLabelHyphenated = hyphenateText(
            sectionSheet.cell(row=j, column=1).value)  # convert spaces in file name to hyphens
        try:
            os.mkdir(f"F:/Valculus/sectionMaterials/{sectionSheet.cell(row=j, column=1).value}")
            with open(
                    f"F:/Valculus/sectionMaterials/{sectionSheet.cell(row=j, column=1).value}/content-{sectionLabelHyphenated}.html",
                    'x') as htmlContent:
                htmlContent.write('')
            with open(
                    f"F:/Valculus/sectionMaterials/{sectionSheet.cell(row=j, column=1).value}/exerciseList-{sectionLabelHyphenated}.html",
                    'x') as htmlExerciseList:
                htmlExerciseList.write('')
        except:
            pass

        try:
            pageCreate.create_all_pages(f'{i}.{j + 1}', f'{topics_list[i][j]}', f'{i}')
        except:
            pass